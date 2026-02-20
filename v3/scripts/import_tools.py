"""
Clean up the Excel data and import into the MakerLAB Inventory AirTable base.

Usage:
  1. Run setup_airtable.py first to create the base
  2. Add AIRTABLE_BASE_ID to your .env file
  3. Run: python import_tools.py

The script will:
  - Read the Excel file from v0/tools.xlsx
  - Clean tool names (trim whitespace)
  - Remove category headers (not real tools)
  - Deduplicate tools (merge by name, track serial numbers)
  - Assign zones and types based on position/name
  - Convert Google Drive URLs to thumbnail format
  - Clear placeholder images
  - Create Tool records in AirTable
  - Create Unit records linked to their Tools
"""

import json
import os
import re
import sys
import time
import urllib.request
import urllib.error

# Requires openpyxl: pip install openpyxl
import openpyxl

API_URL = "https://api.airtable.com/v0"

# The placeholder image ID used for rows 35+
PLACEHOLDER_IMAGE_ID = "1dCH1vh3slhcswtGMNkvmrcXDd7Iqlreb"

# Category headers that aren't real tools
CATEGORY_HEADERS = {
    "Laser Cutter & Engraver",
    "CNC / Milling",
    "FDM Printers",
    "SLA 3D Printer",
}

# Zone assignment rules based on tool name patterns
ZONE_RULES = [
    # 3D Printing
    (r"Ultimaker|Prusa|Bambu|Form \d|Form Cure|Form Wash|Mayku|Formbox", "3D Printing"),
    # Laser Cutting
    (r"Epilog|Trotec|Laser|Bofa.*Fume", "Laser Cutting"),
    # CNC
    (r"Bantam|CNC|Shopbot|ShopBot|Shaper|Roland.*Vinyl|Cricut", "CNC"),
    # Electronics
    (r"BGA|SMD|Solder|Oscilloscope|Rework Station|DREMEL Workstation", "Electronics"),
    # Scanning/VR
    (r"Scanner|Sprout|GoPro|Tripod|Structure Sensor|iPad|Apple Pencil|Meta Quest|VR", "Scanning/VR"),
    # Sewing
    (r"Singer|EverSewn|Sewing|Embroidery", "Sewing"),
    # Large Format
    (r"DesignJet|Plotter|Label Maker|B&W Laser Printer", "Large Format"),
    # Woodshop (default for most hand tools and power tools)
    (r"RYOBI|MILWAUKEE|BOSCH|SKIL|STANLEY|FESTOOL|Saw|Drill|Sander|Hammer|Pliers|Wrench|Chisel|Plane|Rasp|File|Screwdriver|Snips|Cutter|Staple|Clamp|Mallet|Spokeshave|Nailer|Router Plane|Belt Sander|Bandsaw|WAZER|Rockwell|MAKITA|DEWALT|JET|Bench.*Plane|Block Plane|Back Saw|Dozuki|Dead Blow|Scraper|Heat Gun|Shears|Measuring Tape|Socket|Woodworking|Plywood|Cart|Dust Mask|Dust Extractor|Hose Coupler|PVC Hose|Compressor|Sanding Sheet|CALIFORNIA AIR|HUSKY|Hercules", "Woodshop"),
]

# Type assignment rules
TYPE_RULES = [
    # Consumables
    (r"Dust Mask|Sanding Sheet|Replacement Blade|PVC Hose", "Consumable"),
    # Accessories
    (r"Battery|Charger|Hose Coupler|Hose Ring Clamp|Bit Set|Screwdriving Bit|Expansion Kit|Air Manager|Enclosure Bundle|Tripod|Apple Pencil|Drill Bit|Socket.*Bit|Fume Extractor|Dust Extractor|Label Maker", "Accessory"),
    # Machines (large/powered equipment)
    (r"Ultimaker|Prusa|Bambu|Form \d|Form Cure|Form Wash|Mayku|Epilog|Trotec|Bantam|Shopbot|ShopBot|Shaper Origin|Roland|Cricut|Singer|EverSewn|Scanner|Sprout|GoPro|Meta Quest|Oscilloscope|BGA|SMD|Rework Station|DesignJet|Plotter|WAZER|Rockwell|Bandsaw|DREMEL Workstation|Drill Press|Laser Printer|iPad", "Machine"),
    # Tools (default)
    (r".*", "Tool"),
]


def get_config():
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    config = {}
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if "=" in line and not line.startswith("#"):
                    key, val = line.split("=", 1)
                    config[key] = val

    token = config.get("AIRTABLE_API_KEY") or os.environ.get("AIRTABLE_API_KEY")
    base_id = config.get("AIRTABLE_BASE_ID") or os.environ.get("AIRTABLE_BASE_ID")

    if not token:
        print("Error: AIRTABLE_API_KEY not found in .env or environment.")
        sys.exit(1)
    if not base_id:
        print("Error: AIRTABLE_BASE_ID not found in .env or environment.")
        print("Run setup_airtable.py first, then add the base ID to .env")
        sys.exit(1)

    return token, base_id


def api_request(method, path, token, data=None):
    url = f"{API_URL}{path}"
    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(
        url,
        data=body,
        method=method,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
    )
    try:
        resp = urllib.request.urlopen(req)
        return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        error_body = e.read().decode()
        print(f"API Error {e.code}: {error_body}")
        sys.exit(1)


def get_table_ids(token, base_id):
    """Get table IDs by name."""
    data = api_request("GET", f"/meta/bases/{base_id}/tables", token)
    tables = {}
    for table in data["tables"]:
        tables[table["name"]] = table["id"]
    return tables


def clean_image_url(url):
    """Convert Google Drive URLs to thumbnail format, clear placeholders."""
    if not url:
        return None

    url = str(url).strip()

    # Check if it's the placeholder image
    if PLACEHOLDER_IMAGE_ID in url:
        return None

    # Extract Google Drive file ID and convert to thumbnail URL
    match = re.search(r"id=([a-zA-Z0-9_-]+)", url)
    if match:
        file_id = match.group(1)
        return f"https://drive.google.com/thumbnail?id={file_id}&sz=w400"

    return url


def classify(name, rules):
    """Match a tool name against classification rules."""
    for pattern, value in rules:
        if re.search(pattern, name, re.IGNORECASE):
            return value
    return None


def load_serial_numbers():
    """Load serial numbers from the Sections sheet."""
    excel_path = os.path.join(
        os.path.dirname(__file__), "..", "..", "v0", "tools.xlsx"
    )
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb["Sections In Lab"]

    serials = {}  # tool_name -> [serial_numbers]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 3:
            continue
        name = str(row[0]).strip() if row[0] else None
        serial = str(row[2]).strip() if row[2] else None
        if name and serial and not serial.startswith("http") and serial != "None":
            # Skip header rows and non-serial values
            if serial in ("Serial Number", "Serial Number "):
                continue
            serials.setdefault(name, []).append(serial)

    return serials


def load_and_clean_tools():
    """Load Excel data, clean it, and return deduplicated tools."""
    excel_path = os.path.join(
        os.path.dirname(__file__), "..", "..", "v0", "tools.xlsx"
    )
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb["Studio 101 Tools & Equipment"]

    serial_numbers = load_serial_numbers()

    # First pass: collect all rows
    raw_tools = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if row[0] else None
        image_url = str(row[1]).strip() if row[1] else None
        description = str(row[2]).strip() if row[2] else None

        if not name:
            continue

        # Skip category headers
        if name in CATEGORY_HEADERS:
            continue

        raw_tools.append({
            "name": name,
            "image_url": clean_image_url(image_url),
            "description": description if description and description != "None" else None,
        })

    # Deduplicate: merge by name, collect serial numbers
    seen = {}
    tools = []
    for t in raw_tools:
        key = t["name"]
        if key not in seen:
            serials = serial_numbers.get(t["name"], [])
            tool = {
                "name": t["name"],
                "description": t["description"],
                "image_url": t["image_url"],
                "zone": classify(t["name"], ZONE_RULES),
                "type": classify(t["name"], TYPE_RULES),
                "serials": serials,
                # Only count multiple units if we have multiple serial numbers
                # Otherwise assume duplicates in Excel are data entry errors
                "count": max(1, len(serials)),
            }
            seen[key] = len(tools)
            tools.append(tool)
        else:
            idx = seen[key]
            # Fill in description if this duplicate has one
            if t["description"] and not tools[idx]["description"]:
                tools[idx]["description"] = t["description"]
            # Fill in image if this duplicate has one
            if t["image_url"] and not tools[idx]["image_url"]:
                tools[idx]["image_url"] = t["image_url"]

    return tools


def create_tool_records(token, base_id, tools_table_id, tools):
    """Create Tool records in AirTable. Returns list of (tool_name, record_id)."""
    created = []

    # AirTable allows up to 10 records per request
    for i in range(0, len(tools), 10):
        batch = tools[i : i + 10]
        records = []
        for tool in batch:
            fields = {"name": tool["name"]}
            if tool["description"]:
                fields["description"] = tool["description"]
            if tool["image_url"]:
                fields["image_url"] = tool["image_url"]
            if tool["zone"]:
                fields["zone"] = tool["zone"]
            if tool["type"]:
                fields["type"] = tool["type"]
            records.append({"fields": fields})

        result = api_request(
            "POST",
            f"/{base_id}/{tools_table_id}",
            token,
            {"records": records},
        )

        for j, record in enumerate(result["records"]):
            tool = batch[j]
            created.append((tool["name"], record["id"], tool))
            print(f"  Created tool: {tool['name']} ({record['id']})")

        # Rate limit: 5 requests per second
        time.sleep(0.25)

    return created


def create_unit_records(token, base_id, units_table_id, tool_records):
    """Create Unit records linked to Tools."""
    units = []

    for tool_name, tool_record_id, tool in tool_records:
        if tool["serials"]:
            # One unit per serial number
            for idx, serial in enumerate(tool["serials"]):
                label = f"#{idx + 1}" if len(tool["serials"]) > 1 else None
                units.append({
                    "fields": {
                        "tool": [tool_record_id],
                        "serial_number": serial,
                        "status": "Available",
                        **({"label": label} if label else {}),
                    }
                })
        else:
            # Create units based on duplicate count
            for idx in range(tool["count"]):
                label = f"#{idx + 1}" if tool["count"] > 1 else None
                units.append({
                    "fields": {
                        "tool": [tool_record_id],
                        "status": "Available",
                        **({"label": label} if label else {}),
                    }
                })

    # Batch create units
    created_count = 0
    for i in range(0, len(units), 10):
        batch = units[i : i + 10]
        result = api_request(
            "POST",
            f"/{base_id}/{units_table_id}",
            token,
            {"records": batch},
        )
        created_count += len(result["records"])
        time.sleep(0.25)

    return created_count


def main():
    token, base_id = get_config()

    # Get table IDs
    print("Fetching table IDs...")
    tables = get_table_ids(token, base_id)
    tools_table_id = tables.get("Tools")
    units_table_id = tables.get("Units")

    if not tools_table_id or not units_table_id:
        print("Error: Tools or Units table not found. Run setup_airtable.py first.")
        sys.exit(1)

    # Load and clean data
    print("Loading and cleaning Excel data...")
    tools = load_and_clean_tools()

    print(f"  Found {len(tools)} unique tools (after dedup and cleanup)")
    print()

    # Summary before import
    zones = {}
    types = {}
    for t in tools:
        zones[t["zone"]] = zones.get(t["zone"], 0) + 1
        types[t["type"]] = types.get(t["type"], 0) + 1

    print("By zone:")
    for zone, count in sorted(zones.items(), key=lambda x: (x[0] is None, x[0])):
        print(f"  {zone or 'UNCLASSIFIED'}: {count}")
    print()
    print("By type:")
    for typ, count in sorted(types.items(), key=lambda x: (x[0] is None, x[0])):
        print(f"  {typ or 'UNCLASSIFIED'}: {count}")
    print()

    confirm = input("Proceed with import? (y/n): ")
    if confirm.lower() != "y":
        print("Aborted.")
        sys.exit(0)

    # Create tool records
    print()
    print("Creating Tool records...")
    tool_records = create_tool_records(token, base_id, tools_table_id, tools)
    print(f"  Created {len(tool_records)} tool records")

    # Create unit records
    print()
    print("Creating Unit records...")
    unit_count = create_unit_records(token, base_id, units_table_id, tool_records)
    print(f"  Created {unit_count} unit records")

    print()
    print("Import complete!")
    print(f"  Tools: {len(tool_records)}")
    print(f"  Units: {unit_count}")
    print(f"  Base URL: https://airtable.com/{base_id}")


if __name__ == "__main__":
    main()
