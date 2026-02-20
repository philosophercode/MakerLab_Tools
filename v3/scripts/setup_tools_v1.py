"""
Create the Tools_v1 table from the MakerLAB metadata form responses.

Maps all 17 form columns to AirTable fields, creates the table,
and imports the data.

Usage:
  python setup_tools_v1.py
"""

import json
import os
import re
import sys
import time
import urllib.request
import urllib.error

import openpyxl

API_URL = "https://api.airtable.com/v0"

FORM_XLSX = os.path.join(
    os.path.dirname(__file__),
    "..",
    "..",
    "MakerLAB Tools & Equipment Meta Data Generator (Responses).xlsx",
)


def get_config():
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    config = {}
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if "=" in line and not line.startswith("#"):
                    key, val = line.split("=", 1)
                    config[key] = val

    token = config.get("AIRTABLE_API_KEY") or os.environ.get("AIRTABLE_API_KEY")
    base_id = config.get("AIRTABLE_BASE_ID") or os.environ.get("AIRTABLE_BASE_ID")

    if not token or not base_id:
        print("Error: AIRTABLE_API_KEY and AIRTABLE_BASE_ID required in .env")
        sys.exit(1)

    return token, base_id


def api_request(method, path, token, data=None):
    url = f"{API_URL}{path}"
    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(
        url,
        data=body,
        method=method,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
    )
    try:
        resp = urllib.request.urlopen(req)
        return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        error_body = e.read().decode()
        print(f"API Error {e.code}: {error_body}")
        sys.exit(1)


def create_table(token, base_id):
    """Create Tools_v1 table with all fields from the form."""

    # Collect unique categories and zones from form data
    categories, zones = load_select_options()

    payload = {
        "name": "Tools_v1",
        "fields": [
            # Primary field
            {"name": "name", "type": "singleLineText"},
            # Core info
            {"name": "description", "type": "multilineText"},
            {
                "name": "category",
                "type": "singleSelect",
                "options": {"choices": [{"name": c} for c in sorted(categories)]},
            },
            {
                "name": "location_zone",
                "type": "singleSelect",
                "options": {"choices": [{"name": z} for z in sorted(zones)]},
            },
            {"name": "map_tag", "type": "singleLineText"},
            # Image - filename from form (actual images uploaded separately)
            {"name": "image_filename", "type": "singleLineText"},
            {"name": "image_attachments", "type": "multipleAttachments"},
            # Materials & access
            {"name": "materials", "type": "singleLineText"},
            {
                "name": "authorized_only",
                "type": "singleSelect",
                "options": {
                    "choices": [{"name": "YES"}, {"name": "NO"}]
                },
            },
            {"name": "training_required", "type": "singleLineText"},
            {"name": "use_restrictions", "type": "multilineText"},
            # Documentation URLs
            {"name": "safety_doc_url", "type": "url"},
            {"name": "sop_url", "type": "url"},
            {"name": "video_url", "type": "url"},
            # Safety
            {"name": "emergency_stop", "type": "multilineText"},
            {"name": "ppe_required", "type": "singleLineText"},
            # Search & discovery
            {"name": "tags", "type": "singleLineText"},
            # Attachments for manuals (staff uploads later)
            {"name": "manual_attachments", "type": "multipleAttachments"},
        ],
    }

    print("Creating Tools_v1 table...")
    result = api_request("POST", f"/meta/bases/{base_id}/tables", token, payload)
    table_id = result["id"]
    print(f"  Table ID: {table_id}")
    return table_id


def load_select_options():
    """Pre-scan form data to collect unique categories and zones."""
    wb = openpyxl.load_workbook(FORM_XLSX, read_only=True)
    ws = wb["Form Responses 1"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    categories = set()
    zones = set()
    for r in rows:
        r = list(r) + [None] * (17 - len(r))
        if r[2] and str(r[2]).strip():
            categories.add(str(r[2]).strip())
        if r[3] and str(r[3]).strip():
            zones.add(str(r[3]).strip())

    return categories, zones


def clean_url(val):
    """Clean a URL field - return None if not a valid URL."""
    if not val:
        return None
    val = str(val).strip()
    if val.lower() in ("none", "n/a", "", "none "):
        return None
    # Must start with http or www
    if val.startswith("http://") or val.startswith("https://"):
        return val
    if val.startswith("www."):
        return f"https://{val}"
    # Could be a Google Docs URL missing the protocol
    if "docs.google.com" in val or "drive.google.com" in val:
        return f"https://{val}"
    return None


def clean_text(val):
    """Clean a text field."""
    if not val:
        return None
    val = str(val).strip()
    if val.lower() in ("none", "n/a", ""):
        return None
    return val


def load_form_data():
    """Load and clean form responses."""
    wb = openpyxl.load_workbook(FORM_XLSX, read_only=True)
    ws = wb["Form Responses 1"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    tools = []
    seen_names = set()

    for r in rows:
        r = list(r) + [None] * (17 - len(r))

        name = clean_text(r[1])
        if not name:
            continue

        # Deduplicate by name - keep the most recent (later row)
        if name in seen_names:
            # Find and replace the earlier entry
            for i, t in enumerate(tools):
                if t is not None and t["name"] == name:
                    tools[i] = None  # mark for removal
                    break
        seen_names.add(name)

        tool = {
            "name": name,
            "description": clean_text(r[5]),
            "category": clean_text(r[2]),
            "location_zone": clean_text(r[3]),
            "map_tag": clean_text(r[4]),
            "image_filename": clean_text(r[6]),
            "materials": clean_text(r[7]),
            "authorized_only": clean_text(r[8]),
            "training_required": clean_text(r[9]),
            "use_restrictions": clean_text(r[10]),
            "safety_doc_url": clean_url(r[11]),
            "sop_url": clean_url(r[12]),
            "emergency_stop": clean_text(r[13]),
            "ppe_required": clean_text(r[14]),
            "video_url": clean_url(r[15]),
            "tags": clean_text(r[16]),
        }
        tools.append(tool)

    # Remove None entries from dedup
    tools = [t for t in tools if t is not None]
    return tools


def import_tools(token, base_id, table_id, tools):
    """Import tool records to AirTable."""
    created = 0

    for i in range(0, len(tools), 10):
        batch = tools[i: i + 10]
        records = []
        for tool in batch:
            fields = {"name": tool["name"]}
            if tool["description"]:
                fields["description"] = tool["description"]
            if tool["category"]:
                fields["category"] = tool["category"]
            if tool["location_zone"]:
                fields["location_zone"] = tool["location_zone"]
            if tool["map_tag"]:
                fields["map_tag"] = tool["map_tag"]
            if tool["image_filename"]:
                fields["image_filename"] = tool["image_filename"]
            if tool["materials"]:
                fields["materials"] = tool["materials"]
            if tool["authorized_only"]:
                auth = tool["authorized_only"].upper()
                if auth in ("YES", "NO"):
                    fields["authorized_only"] = auth
            if tool["training_required"]:
                fields["training_required"] = tool["training_required"]
            if tool["use_restrictions"]:
                fields["use_restrictions"] = tool["use_restrictions"]
            if tool["safety_doc_url"]:
                fields["safety_doc_url"] = tool["safety_doc_url"]
            if tool["sop_url"]:
                fields["sop_url"] = tool["sop_url"]
            if tool["emergency_stop"]:
                fields["emergency_stop"] = tool["emergency_stop"]
            if tool["ppe_required"]:
                fields["ppe_required"] = tool["ppe_required"]
            if tool["video_url"]:
                fields["video_url"] = tool["video_url"]
            if tool["tags"]:
                fields["tags"] = tool["tags"]
            records.append({"fields": fields})

        result = api_request(
            "POST", f"/{base_id}/{table_id}", token, {"records": records}
        )
        created += len(result["records"])

        for j, rec in enumerate(result["records"]):
            print(f"  {batch[j]['name']}")

        time.sleep(0.25)

    return created


def main():
    token, base_id = get_config()

    # Load and clean form data
    print("Loading form responses...")
    tools = load_form_data()
    print(f"  {len(tools)} unique tools")

    # Summary
    filled = {}
    for key in tools[0].keys():
        filled[key] = sum(1 for t in tools if t[key])
    print("\nField coverage:")
    for key, count in filled.items():
        print(f"  {key:25s} {count:3d}/{len(tools)}")

    print()

    # Create table
    table_id = create_table(token, base_id)

    # Import
    print(f"\nImporting {len(tools)} tools...")
    count = import_tools(token, base_id, table_id, tools)
    print(f"\nDone! Created {count} records.")
    print(f"  Table: Tools_v1 ({table_id})")
    print(f"  Base URL: https://airtable.com/{base_id}")


if __name__ == "__main__":
    main()
